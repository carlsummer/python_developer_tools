# !/usr/bin/env python
# -- coding: utf-8 --
# @Author zengxiaohui
# Datatime:6/4/2021 5:11 PM
# @File:xxxxx
from python_developer_tools.files.common import get_filename_suf_pix
from python_developer_tools.files.image_utils import get_all_images_by_dir
import openpyxl
from openpyxl.styles import Side, Border, colors
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl.drawing.image import Image


def border(t_border, b_border, l_border, r_border):
    """
    :param t_border: side style('dashDot','dashDotDot', 'dashed','dotted',
                        'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                        'mediumDashed', 'slantDashDot', 'thick', 'thin')
    :param b_border: side style
    :param l_border: side style
    :param r_border: side style
    :return:
    """
    border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                    bottom=Side(border_style=b_border, color=colors.BLACK),
                    left=Side(border_style=l_border, color=colors.BLACK),
                    right=Side(border_style=r_border, color=colors.BLACK))
    return border

def execl_template_modify():
    """读取模板xlsx，然后给模板添加内容"""
    dir = r"\\10.20.200.170\data\ext\PVDefectData\haining\qietudedatu(qietudedatu)\waiguantupian_BigImgXml"
    save_excel_path = r"C:\Users\zengxh\Desktop\images.xlsx"

    exl = openpyxl.load_workbook("empty.xlsx")
    sht = exl.get_sheet_by_name("Sheet1")
    Color = ['BDD7EE', 'F4B084']

    imagePaths = get_all_images_by_dir(dir)
    for idx, imagePath in enumerate(imagePaths):
        filename, _, _, _ = get_filename_suf_pix(imagePath)
        sht.append([filename, "", ""])

        col_a = sht["A"]
        col_b = sht["B"]
        col_c = sht["C"]
        col_a[idx + 1].border = border('thin', 'thin', 'thin', 'thin')

        col_b[idx + 1].fill = PatternFill('solid', fgColor=Color[0])
        col_b[idx + 1].border = border('thin', 'thin', 'thin', 'thin')

        col_c[idx + 1].fill = PatternFill('solid', fgColor=Color[1])
        col_c[idx + 1].border = border('thin', 'thin', 'thin', 'thin')
    exl.save(save_excel_path)
    # print("保存excel文件:", save_excel_path)
